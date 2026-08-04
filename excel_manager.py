"""Leitura e preenchimento de arquivos XLSX sem tocar no original."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from hashlib import sha256
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

from utils import normalize_reference


@dataclass
class FillResult:
    workbook_bytes: bytes
    updated_count: int
    missing_references: list[str]


def workbook_fingerprint(source_bytes: bytes) -> str:
    return sha256(source_bytes).hexdigest()


def model_fingerprint(sheet_name: str, headers: list[dict[str, Any]]) -> str:
    """Stable key: adding/removing units does not invalidate a saved model."""
    layout = "|".join(f"{item['column']}:{item['label'].split(' — ', 1)[-1]}" for item in headers)
    return sha256(f"{sheet_name}|{layout}".encode()).hexdigest()


def _load(source_bytes: bytes):
    return load_workbook(BytesIO(source_bytes), data_only=False, keep_links=True)


def get_sheet_names(source_bytes: bytes) -> list[str]:
    return _load(source_bytes).sheetnames


def detect_header_candidates(source_bytes: bytes, sheet_name: str, scan_rows: int = 30, include_blank: bool = False) -> list[dict[str, Any]]:
    """Return the most likely header row, represented as selectable labels."""
    ws = _load(source_bytes)[sheet_name]
    limit = min(ws.max_row, scan_rows)
    best_row, best_cells = 1, []
    for row in ws.iter_rows(min_row=1, max_row=limit):
        cells = [cell for cell in row if cell.value is not None and str(cell.value).strip()]
        # In modelos comuns, uma primeira linha traz título/data e a linha
        # seguinte traz os cabeçalhos. Em caso de empate, a última linha é a
        # mais provável (e evita escolher o título superior).
        if len(cells) >= len(best_cells):
            best_row, best_cells = row[0].row, cells
    if not best_cells:
        return []
    result = []
    for cell in ws[best_row]:
        value = cell.value
        if value is None and not include_blank:
            continue
        name = str(value).strip() if value is not None else "Sem título"
        result.append({"label": f"{get_column_letter(cell.column)} — {name}", "column": cell.column, "row": best_row})
    return result


def read_sheet_map(source_bytes: bytes, sheet_name: str, max_rows: int = 40) -> list[dict[str, Any]]:
    ws = _load(source_bytes)[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows)):
        values = [str(cell.value)[:80] if cell.value is not None else "" for cell in row]
        if any(values):
            rows.append({"Linha": row[0].row, **{get_column_letter(cell.column): value for cell, value in zip(row, values) if value}})
    return rows


def read_units(source_bytes: bytes, sheet_name: str, unit_column: int, code_column: int, start_row: int) -> list[dict[str, Any]]:
    ws = _load(source_bytes)[sheet_name]
    units = []
    for row in range(start_row, ws.max_row + 1):
        unit, code = ws.cell(row, unit_column).value, ws.cell(row, code_column).value
        if unit is not None and code is not None and str(code).strip():
            units.append({"Unidade": str(unit).strip(), "Código": str(code).strip(), "Ativa": True})
    return units


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source, target = ws.cell(source_row, column), ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def apply_unit_catalog(ws, unit_column: int, code_column: int, destination_column: int, start_row: int,
                       units: list[dict[str, Any]]) -> None:
    """Append new active units without moving or restyling existing records."""
    active = [item for item in units if item.get("Ativa", True) and str(item.get("Código", "")).strip()]
    # As larguras são propriedades da coluna: guardamos e restauramos para
    # assegurar que um cadastro nunca modifica o desenho do modelo.
    widths = {letter: dimension.width for letter, dimension in ws.column_dimensions.items()}
    existing_rows = [row for row in range(start_row, ws.max_row + 1)
                     if ws.cell(row, unit_column).value is not None and ws.cell(row, code_column).value is not None]
    existing_codes = {normalize_reference(ws.cell(row, code_column).value) for row in existing_rows}
    style_row = existing_rows[-1] if existing_rows else start_row
    new_units = [item for item in active if normalize_reference(item.get("Código", "")) not in existing_codes]
    for offset, item in enumerate(new_units, start=1):
        target_row = style_row + offset
        _copy_row_style(ws, style_row, target_row)
        ws.cell(target_row, unit_column).value = item.get("Unidade", "")
        ws.cell(target_row, code_column).value = item.get("Código", "")
        ws.cell(target_row, destination_column).value = None
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def build_filled_workbook(source_bytes: bytes, sheet_name: str, reference_column: int, destination_column: int,
                          start_row: int, entries: list[dict[str, Any]], date_cell: str | None = None,
                          date_value: date | None = None, unit_column: int | None = None,
                          units: list[dict[str, Any]] | None = None) -> FillResult:
    """Write only chosen destination cells in an in-memory copy of the workbook."""
    if reference_column == destination_column:
        raise ValueError("A coluna de referência e a coluna de destino devem ser diferentes.")
    wb = _load(source_bytes)
    ws = wb[sheet_name]
    if units is not None and unit_column is not None:
        apply_unit_catalog(ws, unit_column, reference_column, destination_column, start_row, units)
    entry_values: dict[str, Any] = {}
    for entry in entries:
        reference = normalize_reference(entry.get("Referência", ""))
        if reference:
            entry_values[reference] = entry.get("Valor")
    if not entry_values:
        raise ValueError("Não há dados válidos para preencher.")

    found: set[str] = set()
    updated = 0
    for row in range(start_row, ws.max_row + 1):
        key = normalize_reference(ws.cell(row, reference_column).value)
        if key in entry_values:
            ws.cell(row, destination_column).value = entry_values[key]
            found.add(key)
            updated += 1
    if date_cell:
        try:
            ws[date_cell] = date_value
        except ValueError as exc:
            raise ValueError("A célula da data deve ter formato como C1.") from exc

    output = BytesIO()
    wb.save(output)
    missing = [key for key in entry_values if key not in found]
    return FillResult(output.getvalue(), updated, missing)


def find_matching_references(source_bytes: bytes, sheet_name: str, reference_column: int,
                              start_row: int, entries: list[dict[str, Any]]) -> list[str]:
    """Return references from the preview that exist in the configured column."""
    wb = _load(source_bytes)
    ws = wb[sheet_name]
    requested = {normalize_reference(row.get("Referência", "")) for row in entries}
    requested.discard("")
    matches: list[str] = []
    for row in range(start_row, ws.max_row + 1):
        value = normalize_reference(ws.cell(row, reference_column).value)
        if value in requested and value not in matches:
            matches.append(value)
    return matches
